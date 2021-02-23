import os
import re
import xlrd
import xlwt

import openpyxl as opl

import win32com.client as win32

import pandas as pd

count = 0

rootdir = 'D:\\XJTLU UPD DCW\\UPD Data Property\\丹东\\建筑形态\\低层（0-3层）' #change here

#---------------------------------------------------------------------------------------------------

def get_files(path=rootdir,rule='.xls'):
    all = []
    for fpathe,dirs,fs in os.walk(path):   # os.walk是获取所有的目录
        for f in fs:
            filename = os.path.join(fpathe,f)
            if filename.endswith(rule):  # 判断是否是"xxx"结尾
                all.append(filename)
    return all

#---------------------------------------------------------------------------------------------------
def delete_re_rows():
    global count
    b = get_files()
    for pa in b:
        
        #print i
        #list = os.listdir(rootdir) #列出文件夹下所有的目录与文件
        #for i in range(0,len(list)):
        #path = os.path.join(rootdir,list[i])
        #if os.path.isfile(path):#你想对文件的操作
        

        data= xlrd.open_workbook(pa)

        nums = len(data.sheets())

        sheet1 = data.sheets()[0]

        nrows = sheet1.nrows
        ncols = sheet1.ncols

        rows_get = []

        for i in range(nrows):
            A0 = sheet1.cell(i,0).value
            A0 = A0.strip()

            if i<2:
                rows_get.append(i)
            else:
                p = r'[\u4e00-\u9fa5]'

                pattern = re.compile(p)

                try:
                    chk_first = re.findall(pattern,A0)[0]

                    if A0[0:4] == '当页汇总':
                        pass
                    else:
                        rows_get.append(i)
                except:
                    continue

        workbook = xlwt.Workbook('ascii')
        sheet_w = workbook.add_sheet('write')

        wx = 0

        for x in rows_get:
            for y in range(ncols):
                sheet_w.write(wx, y, sheet1.cell(x,y).value)
            wx += 1
        workbook.save(str(pa))
        count+=1
        print(str(count)+' success')

#---------------------------------------------------------------------------------------------------


def get_filesX(path=rootdir,rule='.xlsx'):
    all = []
    for fpathe,dirs,fs in os.walk(path):   # os.walk是获取所有的目录
        for f in fs:
            filename = os.path.join(fpathe,f)
            if filename.endswith(rule):  # 判断是否是"xxx"结尾
                all.append(filename)
    return all

#---------------------------------------------------------------------------------------------------


def renameXLSX():
    b = get_files()
    for pa in b:
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        wb = excel.Workbooks.Open(pa)
        wb.SaveAs(pa+'x', FileFormat = 51)    #FileFormat = 51 is for .xlsx extension
        wb.Close()                               #FileFormat = 56 is for .xls extension
        
#---------------------------------------------------------------------------------------------------

def addColumn():
    all = get_filesX()
    for i in all:
        dir = []
        dir = i.split('\\')
        wb = opl.load_workbook(i)
        ws = wb.worksheets[0]
        cnt = 0
        while cnt < 4:
            ws.insert_cols(1,1)
            for index, row in enumerate(ws.rows):
                if index == 0:
                    if cnt == 0:
                        row[0].value = '日期'
                    elif cnt == 1:
                        row[0].value = '类型'
                    elif cnt == 2:
                        row[0].value = '分类标准'
                    elif cnt == 3:
                        row[0].value = '城市'
                else:
                    if cnt == 0:
                        row[0].value = dir[6][0:-4]
                    elif cnt == 1:
                        row[0].value = dir[5]
                    elif cnt == 2:
                        row[0].value = dir[4]
                    elif cnt == 3:
                        row[0].value = dir[3]
                    
            wb.save(i)
            cnt+=1

#---------------------------------------------------------------------------------------------------

def deleteXls():
    b = get_files()
    cnt = 1
    for i in b:
        try:
            os.remove(i)
            print(str(cnt)+'delete success')
        except:
            print(str(cnt)+'delete failed')
            continue
        cnt+=1

#---------------------------------------------------------------------------------------------------

def merge():
    list = []
    all = get_filesX()
    cnt = 0
    indexCnt = 1
    for i in all:
        if cnt < 1:
            data = pd.read_excel(i,dtype=object)
        else:
            data = pd.read_excel(i,dtype=object,skiprows=0)
        list.append(data)
        print(str(indexCnt)+'merged')
        indexCnt += 1
        cnt += 1
    pd.concat(list).to_excel('D:\\XJTLU UPD DCW\\UPD Data Property\\舟山\\舟山.xlsx', index = False) #change here

#---------------------------------------------------------------------------------------------------


if __name__ == "__main__":
    delete_re_rows()
    #renameXLSX()
    #addColumn()
    #deleteXls()
    #merge()
    